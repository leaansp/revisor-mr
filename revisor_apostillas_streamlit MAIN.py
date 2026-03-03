#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
REVISOR AUTOMÁTICO DE DOCUMENTOS PARA APOSTILLAS
Versión 4.0 – Con lógica de vinculación IF ↔ CE (Partidas de nacimiento GCABA)

MEJORAS EN ESTA VERSIÓN:
- Detecta automáticamente archivos IF y CE por nombre
- Vincula cada CE con su IF correspondiente verificando el número referenciado
- Combina ambos PDFs para análisis unificado con Claude
- La firma que importa es la del CE (no la del IF)
- Alerta si falta alguno de los dos archivos del par
- Fix: Detecta formato "26 de febrero de 2026" correctamente
- Fix: Prompt 100% dinámico (no hardcoded para ningún año específico)
"""

import os
import base64
import json
import re
import io
from datetime import datetime, timedelta
import streamlit as st
import anthropic
import pandas as pd
from PyPDF2 import PdfReader, PdfWriter
import pdfplumber
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from pdf2image import convert_from_bytes
import numpy as np

st.set_page_config(page_title="Revisor de Apostillas", page_icon="📄", layout="wide")

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Geist:wght@300;400;500;600&family=Geist+Mono:wght@400;500&display=swap');

:root {
    --bg:          #0E0E10;
    --surface:     #18181B;
    --surface2:    #232326;
    --border:      rgba(255,255,255,0.07);
    --accent:      #4F8EF7;
    --accent-dim:  rgba(79,142,247,0.12);
    --text:        #F0F0F2;
    --text-muted:  #7A7A85;
}

html, body, [class*="css"] { font-family: 'Geist', sans-serif !important; }

.stApp { background-color: var(--bg) !important; }
.main .block-container { padding-top: 2rem; padding-bottom: 3rem; max-width: 960px; }

[data-testid="stSidebar"] {
    background-color: var(--surface) !important;
    border-right: 1px solid var(--border) !important;
}
[data-testid="stSidebar"] > div:first-child { padding-top: 1.8rem; }
[data-testid="stSidebar"] * { color: var(--text) !important; }
[data-testid="stSidebar"] hr { border-color: var(--border) !important; margin: 1rem 0 !important; }
[data-testid="stSidebar"] .stTextInput input {
    background-color: var(--surface2) !important;
    border: 1px solid var(--border) !important;
    color: var(--text) !important;
    border-radius: 8px !important;
    font-family: 'Geist Mono', monospace !important;
    font-size: 11.5px !important;
    caret-color: var(--accent);
}
[data-testid="stSidebar"] .stTextInput input:focus {
    border-color: var(--accent) !important;
    box-shadow: 0 0 0 3px var(--accent-dim) !important;
}

.sidebar-label {
    font-size: 10px !important;
    letter-spacing: 0.1em;
    text-transform: uppercase;
    color: var(--text-muted) !important;
    font-weight: 500;
    display: block;
}
.sidebar-title { font-size: 1rem; font-weight: 600; color: var(--text) !important; letter-spacing: -0.02em; margin: 0.2rem 0 0 0; }
.sidebar-sub   { font-size: 11px; color: var(--text-muted) !important; margin-top: 0.15rem; letter-spacing: 0.03em; }

.pill {
    display: inline-flex; align-items: center; gap: 5px;
    font-family: 'Geist Mono', monospace; font-size: 10px; font-weight: 500;
    letter-spacing: 0.05em; color: var(--accent); background: var(--accent-dim);
    padding: 3px 10px; border-radius: 100px; margin-bottom: 1rem;
    border: 1px solid rgba(79,142,247,0.2);
}
.pill-dot {
    width: 6px; height: 6px; border-radius: 50%; background: var(--accent);
    display: inline-block;
    animation: pulse-dot 2s ease-in-out infinite;
}
@keyframes pulse-dot {
    0%, 100% { opacity: 1; transform: scale(1); }
    50%       { opacity: 0.4; transform: scale(0.7); }
}

.main-header { padding: 0 0 2rem 0; border-bottom: 1px solid var(--border); margin-bottom: 2.2rem; }
.main-header h1 { font-size: 1.75rem; font-weight: 600; color: var(--text); letter-spacing: -0.04em; margin: 0 0 0.35rem 0; line-height: 1.15; }
.main-header p  { font-size: 0.875rem; color: var(--text-muted); margin: 0; font-weight: 400; line-height: 1.5; }

[data-testid="metric-container"] {
    background: var(--surface) !important;
    border: 1px solid var(--border) !important;
    border-radius: 12px !important;
    padding: 1.1rem 1.3rem !important;
    transition: border-color 0.2s;
}
[data-testid="metric-container"]:hover { border-color: rgba(255,255,255,0.13) !important; }
[data-testid="metric-container"] label {
    font-size: 10px !important; letter-spacing: 0.09em !important;
    text-transform: uppercase !important; color: var(--text-muted) !important; font-weight: 500 !important;
}
[data-testid="metric-container"] [data-testid="stMetricValue"] {
    font-size: 2.1rem !important; font-weight: 600 !important;
    color: var(--text) !important; letter-spacing: -0.03em; line-height: 1;
}

.stButton > button[kind="primary"] {
    background: var(--accent) !important; color: #fff !important;
    border: none !important; border-radius: 9px !important;
    font-family: 'Geist', sans-serif !important; font-weight: 500 !important;
    font-size: 0.875rem !important; padding: 0.6rem 1.6rem !important;
    transition: all 0.15s ease !important;
    box-shadow: 0 0 20px rgba(79,142,247,0.25);
}
.stButton > button[kind="primary"]:hover {
    background: #6FA3F9 !important;
    box-shadow: 0 0 28px rgba(79,142,247,0.4) !important;
    transform: translateY(-1px);
}

.stDownloadButton > button {
    background: var(--surface2) !important; border: 1px solid var(--border) !important;
    color: var(--text) !important; border-radius: 9px !important;
    font-family: 'Geist', sans-serif !important; font-weight: 500 !important;
    font-size: 0.85rem !important; padding: 0.5rem 1.3rem !important;
    transition: all 0.15s ease !important;
}
.stDownloadButton > button:hover {
    border-color: var(--accent) !important; color: var(--accent) !important;
    background: var(--accent-dim) !important;
}

[data-testid="stFileUploader"] {
    background: transparent !important;
    border-radius: 0 !important;
    padding: 0 !important;
}
[data-testid="stFileUploader"] label { display: none !important; }

/* Dropzone grande, una sola caja, todo clickeable */
[data-testid="stFileUploaderDropzone"] {
    background: var(--surface) !important;
    border: 1.5px dashed rgba(255,255,255,0.09) !important;
    border-radius: 14px !important;
    padding: 3.5rem 2rem 1.5rem 2rem !important;
    transition: border-color 0.2s ease, background 0.2s ease !important;
    cursor: pointer !important;
    text-align: center !important;
    position: relative !important;
    min-height: 240px !important;
    display: flex !important;
    flex-direction: column !important;
    align-items: center !important;
    justify-content: center !important;
}
[data-testid="stFileUploaderDropzone"]:hover {
    border-color: var(--accent) !important;
    background: rgba(79,142,247,0.05) !important;
}
/* Sin ::before — el ícono se inyecta via st.markdown encima del uploader */
[data-testid="stFileUploaderDropzoneInput"] { cursor: pointer !important; }

/* Ocultar el ícono nativo de nube de Streamlit */
[data-testid="stFileUploaderDropzone"] > div > div:first-child,
[data-testid="stFileUploaderDropzone"] svg,
[data-testid="stFileUploaderDropzone"] img {
    display: none !important;
}
/* El div interno que contiene texto — centrado */
[data-testid="stFileUploaderDropzone"] > div {
    display: flex !important;
    flex-direction: column !important;
    align-items: center !important;
    justify-content: center !important;
    width: 100% !important;
}
/* Texto principal "Drag and drop files here" */
[data-testid="stFileUploaderDropzone"] > div > span {
    font-family: 'Geist', sans-serif !important;
    font-size: 0.9rem !important;
    font-weight: 500 !important;
    color: #F0F0F2 !important;
    letter-spacing: -0.01em !important;
    display: block !important;
    text-align: center !important;
}
/* Subtexto "Limit 200MB per file • PDF" */
[data-testid="stFileUploaderDropzone"] small {
    font-size: 0.76rem !important;
    color: var(--text-muted) !important;
    display: block !important;
    margin-top: 0.2rem !important;
    text-align: center !important;
}
/* Botón Browse files */
[data-testid="stFileUploaderDropzone"] button {
    background: var(--surface2) !important;
    border: 1px solid rgba(255,255,255,0.1) !important;
    color: var(--text) !important;
    border-radius: 8px !important;
    font-family: 'Geist', sans-serif !important;
    font-size: 0.8rem !important;
    font-weight: 500 !important;
    padding: 0.4rem 1.2rem !important;
    margin-top: 0.6rem !important;
    transition: all 0.15s ease !important;
    cursor: pointer !important;
}
[data-testid="stFileUploaderDropzone"] button:hover {
    border-color: var(--accent) !important;
    color: var(--accent) !important;
    background: var(--accent-dim) !important;
}
[data-testid="stFileUploaderFile"] {
    background: var(--surface2) !important;
    border: 1px solid var(--border) !important;
    border-radius: 8px !important;
    padding: 0.5rem 0.8rem !important;
    margin-top: 0.4rem !important;
    font-size: 0.82rem !important;
}
[data-testid="stFileUploaderFile"] * { color: var(--text) !important; }
[data-testid="stFileUploader"] label { display: none !important; }

[data-testid="stProgress"] > div {
    background: var(--surface2) !important; border-radius: 100px;
}
[data-testid="stProgress"] > div > div {
    background: linear-gradient(90deg, var(--accent), #7AB8FF) !important;
    border-radius: 100px; transition: width 0.3s ease;
}

p, span, label, div { color: var(--text); }
h1, h2, h3, h4 { color: var(--text) !important; }

[data-testid="stDataFrame"] {
    border: 1px solid var(--border) !important;
    border-radius: 12px !important; overflow: hidden !important;
    background: var(--surface) !important;
}

.stAlert {
    border-radius: 10px !important; font-size: 0.85rem !important;
    background: var(--surface2) !important;
    border: 1px solid var(--border) !important; color: var(--text) !important;
}

[data-testid="stExpander"] {
    background: var(--surface) !important;
    border: 1px solid var(--border) !important; border-radius: 10px !important;
}
[data-testid="stExpander"] summary { font-size: 0.85rem !important; color: var(--text-muted) !important; }

.section-divider { height: 1px; background: var(--border); margin: 1.8rem 0; }
.results-label {
    font-size: 10px; font-weight: 600; letter-spacing: 0.1em;
    text-transform: uppercase; color: var(--text-muted); margin-bottom: 0.8rem;
}
.footer {
    margin-top: 3.5rem; padding-top: 1.5rem; border-top: 1px solid var(--border);
    text-align: center; color: var(--text-muted); font-size: 11px; letter-spacing: 0.03em;
}
/* ── Tabs ── */
[data-testid="stTabs"] {
    margin-top: 0.5rem;
}
[data-testid="stTabBar"] {
    background: transparent !important;
    border-bottom: 1px solid var(--border) !important;
    gap: 0.25rem;
    padding-bottom: 0;
}
[data-testid="stTabBar"] button {
    background: transparent !important;
    border: none !important;
    border-bottom: 2px solid transparent !important;
    border-radius: 0 !important;
    color: var(--text-muted) !important;
    font-family: 'Geist', sans-serif !important;
    font-size: 0.875rem !important;
    font-weight: 500 !important;
    padding: 0.6rem 1rem !important;
    transition: color 0.15s ease, border-color 0.15s ease !important;
    letter-spacing: -0.01em;
}
[data-testid="stTabBar"] button:hover {
    color: var(--text) !important;
    background: transparent !important;
}
[data-testid="stTabBar"] button[aria-selected="true"] {
    color: var(--text) !important;
    border-bottom-color: var(--accent) !important;
    background: transparent !important;
}
[data-testid="stTabContent"] {
    padding-top: 1.8rem;
}
/* Criteria page styles */
.crit-section {
    background: var(--surface);
    border: 1px solid var(--border);
    border-radius: 14px;
    padding: 1.4rem 1.6rem;
    margin-bottom: 1rem;
}
.crit-section-title {
    font-size: 0.78rem;
    font-weight: 600;
    letter-spacing: 0.08em;
    text-transform: uppercase;
    color: var(--text-muted);
    margin: 0 0 1rem 0;
    padding-bottom: 0.75rem;
    border-bottom: 1px solid var(--border);
    display: flex;
    align-items: center;
    gap: 0.6rem;
}
.crit-icon {
    display: inline-flex;
    align-items: center;
    justify-content: center;
    width: 28px;
    height: 28px;
    border-radius: 7px;
    background: var(--surface2);
    border: 1px solid var(--border);
    flex-shrink: 0;
}
.crit-scope-tag {
    display: inline-block;
    font-size: 10px;
    font-weight: 600;
    letter-spacing: 0.06em;
    text-transform: uppercase;
    color: var(--accent);
    background: var(--accent-dim);
    border: 1px solid rgba(79,142,247,0.2);
    padding: 2px 8px;
    border-radius: 100px;
    margin-bottom: 1rem;
}
.crit-row {
    display: flex;
    align-items: baseline;
    gap: 0.75rem;
    padding: 0.5rem 0;
    border-bottom: 1px solid rgba(255,255,255,0.04);
}
.crit-row:last-child { border-bottom: none; }
.crit-badge {
    flex-shrink: 0;
    font-size: 9px;
    font-weight: 700;
    letter-spacing: 0.08em;
    text-transform: uppercase;
    padding: 2px 7px;
    border-radius: 100px;
    min-width: 68px;
    text-align: center;
}
.badge-ok      { background: rgba(52,199,89,0.12);  color: #34C759; border: 1px solid rgba(52,199,89,0.2); }
.badge-warn    { background: rgba(255,159,10,0.12); color: #FF9F0A; border: 1px solid rgba(255,159,10,0.2); }
.badge-danger  { background: rgba(255,69,58,0.12);  color: #FF453A; border: 1px solid rgba(255,69,58,0.2); }
.crit-text {
    font-size: 0.875rem;
    color: #B0B0BA;
    line-height: 1.5;
}
.crit-note {
    font-size: 0.78rem;
    color: var(--text-muted);
    margin-top: 0.25rem;
    font-style: italic;
}

</style>
""", unsafe_allow_html=True)

# ── SIDEBAR ──────────────────────────────────────────────────────────────────
with st.sidebar:
    dias = ["Lunes","Martes","Miércoles","Jueves","Viernes","Sábado","Domingo"]
    meses = ["enero","febrero","marzo","abril","mayo","junio","julio","agosto","septiembre","octubre","noviembre","diciembre"]
    now = datetime.now()
    fecha_bonita = f"{dias[now.weekday()]} {now.day} de {meses[now.month-1]} de {now.year}"
    st.markdown(f'''
        <div class="pill"><span class="pill-dot"></span> v4.0 · Claude Sonnet</div>
        <p class="sidebar-title">Revisor de Apostillas</p>
        <p class="sidebar-sub">Dirección Técnica Consular</p>
        <p style="font-size:10px; color:#48484A; margin: 0.15rem 0 0 0; line-height:1.5;">Ministerio de Relaciones Exteriores, Comercio Internacional y Culto &middot; Cancillería</p>
        <p style="font-size:11px; color:#7A7A85; margin: 0.7rem 0 0 0; letter-spacing:0.01em;">{fecha_bonita}</p>
    ''', unsafe_allow_html=True)
    st.markdown("---")

    st.markdown('<span class="sidebar-label">API Key de Claude</span>', unsafe_allow_html=True)
    CLAUDE_API_KEY = os.environ.get("ANTHROPIC_API_KEY", "")
    if not CLAUDE_API_KEY:
        CLAUDE_API_KEY = st.text_input(
            label="API Key",
            type="password",
            placeholder="sk-ant-api03-...",
            label_visibility="collapsed"
        )
    else:
        st.markdown('<p style="font-size:12px; color:#34C759;">API Key cargada desde entorno.</p>', unsafe_allow_html=True)

    st.markdown("---")
    st.markdown('''
<div style="display:flex; align-items:center; gap:0.5rem; margin-bottom:0.4rem;">
    <div style="display:inline-flex; align-items:center; justify-content:center;
                width:22px; height:22px; border-radius:6px;
                background:var(--surface2); border:1px solid var(--border); flex-shrink:0;">
        <svg width="11" height="11" viewBox="0 0 24 24" fill="none" xmlns="http://www.w3.org/2000/svg">
            <circle cx="12" cy="12" r="9" stroke="#7A7A85" stroke-width="1.8"/>
            <path d="M9.5 9.5a2.5 2.5 0 015 .5c0 1.5-2.5 2-2.5 3.5" stroke="#7A7A85" stroke-width="1.8" stroke-linecap="round"/>
            <circle cx="12" cy="16.5" r="0.75" fill="#7A7A85"/>
        </svg>
    </div>
    <span class="sidebar-label" style="margin:0;">Cómo funciona</span>
</div>
''', unsafe_allow_html=True)
    st.markdown("""
<p style="font-size:12.5px; line-height:1.7; color:#7A7A85; margin-top:0.4rem;">
Subí uno o más PDFs. El sistema los clasifica como <strong style="color:#F0F0F2;">IF</strong> (documento original) o <strong style="color:#F0F0F2;">CE</strong> (certificado) y los empareja por número de expediente.<br><br>
Los pares IF+CE se analizan en conjunto. Los archivos sueltos se procesan individualmente.<br><br>
El resultado incluye estado de firma digital, vigencia y observaciones.
</p>
""", unsafe_allow_html=True)

    st.markdown("---")
    st.markdown('<p style="font-size:11px; color:#5A5A60; margin:0;">Leandro Spinelli · 2026</p>', unsafe_allow_html=True)

if not CLAUDE_API_KEY:
    st.warning("Ingresá tu API Key en el panel lateral para continuar.")
    st.stop()

# ── HEADER PRINCIPAL ─────────────────────────────────────────────────────────
st.markdown("""
<div class="main-header">
    <div style="
        display: inline-flex; align-items: center; gap: 5px;
        font-family: 'Geist Mono', monospace; font-size: 9px; font-weight: 500;
        letter-spacing: 0.08em; text-transform: uppercase;
        margin-bottom: 0.75rem; flex-wrap: wrap; line-height: 1.8;
    ">
        <span style="color:#3D3D42;">Ministerio de Relaciones Exteriores, Comercio Internacional y Culto</span>
        <span style="color:#4A4A52;">›</span>
        <span style="color:#3D3D42;">Cancillería</span>
        <span style="color:#4A4A52;">›</span>
        <span style="color:#3D3D42;">Dirección General de Asuntos Consulares</span>
        <span style="color:#4A4A52;">›</span>
        <span style="color:#A0A0AA; font-weight:600; font-size:10.5px;">Dirección Técnica Consular</span>
    </div>
    <h1>Revisor Automático de Apostillas</h1>
    <p>Cargá los PDFs — el sistema los clasifica, empareja y valida automáticamente con IA.</p>
</div>
""", unsafe_allow_html=True)

tab_revision, tab_criterios = st.tabs(["Revisión", "Criterios normativos"])

with tab_criterios:
    st.markdown("""
<div class="crit-section">
  <span class="crit-section-title">
    <span class="crit-icon">
      <svg width="14" height="14" viewBox="0 0 24 24" fill="none" xmlns="http://www.w3.org/2000/svg">
        <path d="M9 12h6M9 16h6M7 4H5a2 2 0 00-2 2v14a2 2 0 002 2h14a2 2 0 002-2V6a2 2 0 00-2-2h-2M9 4a2 2 0 002 2h2a2 2 0 002-2M9 4a2 2 0 012-2h2a2 2 0 012 2" stroke="#7A7A85" stroke-width="1.6" stroke-linecap="round" stroke-linejoin="round"/>
      </svg>
    </span>
    Documentos de Estado Civil · CABA
  </span>
  <span class="crit-scope-tag">Solo Registro Civil CABA</span>
  <p style="font-size:0.82rem; color:#7A7A85; margin: 0 0 1rem 0; line-height:1.6;">
    La lógica de vinculación IF + CE aplica exclusivamente a partidas de nacimiento, matrimonio y defunción emitidas por el Registro Civil del Gobierno de la Ciudad de Buenos Aires (GCBA). No aplica a documentos de otras provincias ni de otros organismos.
  </p>
  <div class="crit-row">
    <span class="crit-badge badge-ok">Aprobado</span>
    <div><span class="crit-text">CE referencia correctamente al número IF del documento</span></div>
  </div>
  <div class="crit-row">
    <span class="crit-badge badge-ok">Aprobado</span>
    <div><span class="crit-text">Firma digital del CE presente y válida</span></div>
  </div>
  <div class="crit-row">
    <span class="crit-badge badge-warn">Revisar</span>
    <div><span class="crit-text">CE cargado sin su IF correspondiente</span></div>
  </div>
  <div class="crit-row">
    <span class="crit-badge badge-warn">Revisar</span>
    <div><span class="crit-text">IF cargado sin su CE correspondiente</span></div>
  </div>
  <div class="crit-row">
    <span class="crit-badge badge-warn">Revisar</span>
    <div><span class="crit-text">Firma del CE no detectable automáticamente</span></div>
  </div>
  <div class="crit-row">
    <span class="crit-badge badge-danger">Rechazar</span>
    <div><span class="crit-text">CE no referencia al número IF del documento</span></div>
  </div>
  <div class="crit-row">
    <span class="crit-badge badge-danger">Rechazar</span>
    <div><span class="crit-text">Firma digital del CE ausente</span></div>
  </div>
</div>

<div class="crit-section">
  <span class="crit-section-title">
    <span class="crit-icon">
      <svg width="14" height="14" viewBox="0 0 24 24" fill="none" xmlns="http://www.w3.org/2000/svg">
        <path d="M12 3L5 6.5V11c0 4.1 3.1 7.9 7 8.9 3.9-1 7-4.8 7-8.9V6.5L12 3z" stroke="#7A7A85" stroke-width="1.6" stroke-linecap="round" stroke-linejoin="round"/>
        <path d="M9.5 12l2 2 3.5-3.5" stroke="#7A7A85" stroke-width="1.6" stroke-linecap="round" stroke-linejoin="round"/>
      </svg>
    </span>
    Antecedentes Penales
  </span>
  <div class="crit-row">
    <span class="crit-badge badge-ok">Aprobado</span>
    <div><span class="crit-text">Emitido hace 90 días o menos</span><p class="crit-note">Fecha calculada automáticamente desde la emisión hasta hoy.</p></div>
  </div>
  <div class="crit-row">
    <span class="crit-badge badge-ok">Aprobado</span>
    <div><span class="crit-text">Firma digital presente y válida</span></div>
  </div>
  <div class="crit-row">
    <span class="crit-badge badge-warn">Revisar</span>
    <div><span class="crit-text">Fecha de emisión no detectable o ambigua</span></div>
  </div>
  <div class="crit-row">
    <span class="crit-badge badge-danger">Rechazar</span>
    <div><span class="crit-text">Emitido hace más de 90 días</span><p class="crit-note">Cancillería no acepta antecedentes con más de 90 días de antigüedad.</p></div>
  </div>
  <div class="crit-row">
    <span class="crit-badge badge-danger">Rechazar</span>
    <div><span class="crit-text">Firma digital ausente</span></div>
  </div>
</div>

<div class="crit-section">
  <span class="crit-section-title">
    <span class="crit-icon">
      <svg width="14" height="14" viewBox="0 0 24 24" fill="none" xmlns="http://www.w3.org/2000/svg">
        <path d="M12 2L2 7l10 5 10-5-10-5z" stroke="#7A7A85" stroke-width="1.6" stroke-linecap="round" stroke-linejoin="round"/>
        <path d="M2 17l10 5 10-5M2 12l10 5 10-5" stroke="#7A7A85" stroke-width="1.6" stroke-linecap="round" stroke-linejoin="round"/>
      </svg>
    </span>
    Documentos Educativos · Títulos, Analíticos e Intervención Ministerial
  </span>
  <p style="font-size:0.82rem; color:#7A7A85; margin: 0 0 1rem 0; line-height:1.6;">
    Todo documento educativo (título, analítico, certificado de materias, diploma, historia académica, constancia de estudios, etc.) requiere firma visible de la autoridad educativa <strong style="color:#B0B0BA;">y</strong> la intervención de un funcionario ministerial habilitado. El organismo habilitado depende del año del documento.
  </p>
  <div class="crit-row">
    <span class="crit-badge badge-ok">Aprobado</span>
    <div>
      <span class="crit-text">Documento posterior a 2012 · Sello del Ministerio de Educación confirmado</span>
      <p class="crit-note">Incluye: Ministerio de Capital Humano, Ministerio de Educación e Innovación (CABA), GCBA Legalizaciones — Ministerio de Educación, Dirección General de Cultura y Educación (PBA), DiNIECE.</p>
    </div>
  </div>
  <div class="crit-row">
    <span class="crit-badge badge-ok">Aprobado</span>
    <div>
      <span class="crit-text">Documento anterior a 2012 · Sello del Ministerio del Interior confirmado</span>
      <p class="crit-note">Para documentos emitidos antes de 2012, el organismo habilitado para legalizar documentos educativos era el Ministerio del Interior, no el de Educación.</p>
    </div>
  </div>
  <div class="crit-row">
    <span class="crit-badge badge-warn">Revisar</span>
    <div><span class="crit-text">Sin firma visible detectada</span></div>
  </div>
  <div class="crit-row">
    <span class="crit-badge badge-warn">Revisar</span>
    <div><span class="crit-text">Múltiples firmas con ambigüedad sobre cuál aplica</span></div>
  </div>
  <div class="crit-row">
    <span class="crit-badge badge-warn">Revisar</span>
    <div>
      <span class="crit-text">Sello ministerial presente pero borroso, comprimido o ilegible</span>
      <p class="crit-note">El sello existe visualmente pero no pudo leerse con certeza. Causas frecuentes: PDF enviado por WhatsApp, escaneado de baja calidad, o múltiples sellos superpuestos.</p>
    </div>
  </div>
  <div class="crit-row">
    <span class="crit-badge badge-warn">Revisar</span>
    <div>
      <span class="crit-text">Documento anterior a 2012 · Sello del Ministerio del Interior presente pero firmante no identificado</span>
      <p class="crit-note">El sistema detectó el sello pero no pudo leer el nombre del funcionario habilitado. Verificar manualmente que corresponda al Ministerio del Interior.</p>
    </div>
  </div>
  <div class="crit-row">
    <span class="crit-badge badge-danger">Rechazar</span>
    <div>
      <span class="crit-text">Solo firma de autoridad educativa, sin sello ministerial · Imagen nítida</span>
      <p class="crit-note">El sistema confirmó con certeza que no hay intervención ministerial: imagen clara y los únicos firmantes son rector, decano, director o secretario académico.</p>
    </div>
  </div>
</div>

<div class="crit-section">
  <span class="crit-section-title">
    <span class="crit-icon">
      <svg width="14" height="14" viewBox="0 0 24 24" fill="none" xmlns="http://www.w3.org/2000/svg">
        <circle cx="12" cy="12" r="9" stroke="#7A7A85" stroke-width="1.6"/>
        <path d="M8.5 12.5l2.5 2.5 4.5-5" stroke="#7A7A85" stroke-width="1.6" stroke-linecap="round" stroke-linejoin="round"/>
      </svg>
    </span>
    Criterios generales · Todos los documentos
  </span>
  <div class="crit-row">
    <span class="crit-badge badge-ok">Aprobado</span>
    <div><span class="crit-text">Imagen de calidad alta, clara o nítida</span></div>
  </div>
  <div class="crit-row">
    <span class="crit-badge badge-warn">Revisar</span>
    <div><span class="crit-text">Imagen de calidad baja o borrosa</span></div>
  </div>
  <div class="crit-row">
    <span class="crit-badge badge-warn">Revisar</span>
    <div><span class="crit-text">Documento fotografiado con celular</span></div>
  </div>
  <div class="crit-row">
    <span class="crit-badge badge-danger">Rechazar</span>
    <div><span class="crit-text">Imagen ilegible</span></div>
  </div>
</div>
""", unsafe_allow_html=True)

# =============================================================================
# FIRMA DIGITAL - busca en los mismos lugares que Adobe Reader
# =============================================================================

def verificar_firma_digital(pdf_bytes):
    try:
        reader = PdfReader(io.BytesIO(pdf_bytes))
        firmas = []

        root = reader.trailer.get('/Root', {})
        if hasattr(root, 'get_object'):
            root = root.get_object()

        acroform = root.get('/AcroForm', None)
        if acroform:
            if hasattr(acroform, 'get_object'):
                acroform = acroform.get_object()
            fields = acroform.get('/Fields', [])
            for field_ref in fields:
                try:
                    field = field_ref.get_object()
                    if field.get('/FT') == '/Sig':
                        v = field.get('/V')
                        if v:
                            if hasattr(v, 'get_object'):
                                v = v.get_object()
                            nombre = str(v.get('/Name', ''))
                            razon = str(v.get('/Reason', ''))
                            firmante = nombre or razon or 'Firma digital detectada'
                            if firmante not in firmas:
                                firmas.append(firmante)
                except:
                    continue

        for page in reader.pages:
            try:
                if '/Annots' in page:
                    for annot_ref in page['/Annots']:
                        annot = annot_ref.get_object()
                        if annot.get('/Subtype') == '/Widget' and annot.get('/FT') == '/Sig':
                            v = annot.get('/V')
                            if v:
                                if hasattr(v, 'get_object'):
                                    v = v.get_object()
                                nombre = str(v.get('/Name', 'Firma en página'))
                                if nombre not in firmas:
                                    firmas.append(nombre)
            except:
                continue

        raw = str(reader.trailer)
        if "/Sig" in raw and not firmas:
            firmas.append("Firma digital detectada (certificado no extraíble)")

        return {"tiene_firma": bool(firmas), "cantidad_firmas": len(firmas), "firmantes": firmas}

    except:
        return {"tiene_firma": None, "cantidad_firmas": 0, "firmantes": []}

# =============================================================================
# LÓGICA IF ↔ CE
# =============================================================================

def extraer_clave_if(texto):
    """
    Extrae la clave única de un número IF: tupla (año, numero).
    Ejemplo: "IF-2015-29802485- -DGRC" → ("2015", "29802485")
    Acepta variantes con espacios extra o guiones entre los segmentos.
    """
    match = re.search(r'IF[\s\-_]+(\d{4})[\s\-_]+(\d+)', texto, re.IGNORECASE)
    if match:
        return (match.group(1), match.group(2))
    return None

def extraer_texto_pdf(pdf_bytes):
    """
    Extrae texto de un PDF usando pdfplumber con extract_words().
    extract_words() es mucho más robusto que extract_text() para PDFs de GCABA
    ya que no pierde líneas por problemas de encoding de fuentes.
    """
    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            partes = []
            for page in pdf.pages:
                try:
                    words = page.extract_words()
                    if words:
                        partes.append(" ".join(w["text"] for w in words))
                except:
                    pass
            return " ".join(partes)
    except:
        return ""

def detectar_sello_violeta(pdf_bytes):
    """
    Renderiza las páginas del PDF como imágenes y detecta píxeles violetas/morados.
    El sello rectangular de GCBA Legalizaciones es de tinta violeta/púrpura.
    Retorna True si detecta suficientes píxeles de ese color en cualquier página.
    """
    try:
        pages = convert_from_bytes(pdf_bytes, dpi=100)
        for page in pages:
            arr = np.array(page)
            r, g, b = arr[:,:,0].astype(int), arr[:,:,1].astype(int), arr[:,:,2].astype(int)
            # Violeta/morado: azul y rojo dominantes, verde bajo
            mask = (r > 60) & (r < 220) & (g < 130) & (b > 120) & (b > r) & (b > g + 40)
            if mask.sum() > 300:  # umbral: cluster significativo, no ruido
                return True
    except Exception:
        pass
    return False

def extraer_if_de_bytes_crudos(pdf_bytes):
    """
    Extrae el número IF directamente de los bytes crudos del PDF.
    Funciona incluso cuando el número está en una imagen escaneada,
    porque GEDO lo embebe también como texto en el stream interno del PDF.
    Ejemplo: "IF-2015-29802485- -DGRC" → ("2015", "29802485")
    """
    try:
        raw = pdf_bytes.decode("latin-1", errors="ignore")
        match = re.search(r'IF[-\s]+(\d{4})[-\s]+(\d+)', raw)
        if match:
            return (match.group(1), match.group(2))
    except:
        pass
    return None

def detectar_tipo_por_contenido(pdf_bytes, nombre_archivo=""):
    """
    Clasifica el PDF como CE, IF u OTRO leyendo su CONTENIDO.

    CE → se detecta por "CERTIFICO QUE EL PRESENTE DOCUMENTO" (inequívoco)
         + extrae el número IF referenciado desde el texto seleccionable.

    IF → se detecta por señales de GCABA en el texto.
         Su número IF se extrae de los bytes crudos del PDF (donde GEDO
         lo embebe aunque visualmente esté en la imagen escaneada).
         Esto permite emparejamiento EXACTO con el CE: si los números no
         coinciden, no se emparejan. Sin fallbacks ciegos.

    OTRO → sin señales reconocibles.

    Retorna: ("CE", clave_if_referenciada, texto_debug)
           | ("IF", clave_if_propia, texto_debug)   ← clave puede ser None si no se extrae
           | ("OTRO", None, texto_debug)
    """
    texto_raw = extraer_texto_pdf(pdf_bytes)
    texto_norm = re.sub(' +', ' ', texto_raw.replace('\n', ' ')).strip()
    texto_upper = texto_norm.upper()

    # ── Es CE: frase inequívoca de GCABA ────────────────────────────────────
    if "CERTIFICO QUE EL PRESENTE DOCUMENTO" in texto_upper:
        clave = extraer_clave_if(texto_norm)
        return ("CE", clave, texto_norm)

    # ── Es IF de GCABA: señales de GCABA + número IF con sufijo DGRC ─────────
    # La clave discriminante es el sufijo DGRC (Dirección General del Registro
    # Civil de CABA). Cualquier otro organismo que use GEDO tendrá un sufijo
    # distinto (ej: GDEBA para Provincia de Buenos Aires). Si no hay número IF
    # con DGRC, el documento no es un IF de GCABA aunque use GEDO o tenga
    # "Hoja Adicional de Firmas" — se trata como documento individual (OTRO).
    señales_gcaba = [
        "GOBIERNO DE LA CIUDAD",
        "HOJA ADICIONAL DE FIRMAS",
        "REGISTRO DEL ESTADO CIVIL",
        "GEDO",
    ]
    if any(s in texto_upper for s in señales_gcaba):
        clave_if = extraer_if_de_bytes_crudos(pdf_bytes)
        if clave_if:
            # Verificación extra: el número IF debe venir de un documento DGRC
            raw = pdf_bytes.decode("latin-1", errors="ignore").upper()
            if "DGRC" in raw or "GOBIERNO DE LA CIUDAD" in texto_upper:
                return ("IF", clave_if, texto_norm)
        # Tiene señales de GEDO pero no es GCABA → procesar como individual
        return ("OTRO", None, texto_norm)

    # ── OTRO ─────────────────────────────────────────────────────────────────
    return ("OTRO", None, texto_norm)

def combinar_pdfs(pdf_bytes_lista):
    """Combina múltiples PDFs en uno solo. Retorna los bytes del PDF combinado."""
    writer = PdfWriter()
    for pdf_bytes in pdf_bytes_lista:
        try:
            reader = PdfReader(io.BytesIO(pdf_bytes))
            for page in reader.pages:
                writer.add_page(page)
        except Exception as e:
            st.warning(f"Advertencia al combinar PDF: {e}")
    output = io.BytesIO()
    writer.write(output)
    output.seek(0)
    return output.read()

# =============================================================================
# UTILIDADES
# =============================================================================

def pdf_a_base64(pdf_bytes):
    return base64.b64encode(pdf_bytes).decode('utf-8')

def calcular_dias_desde_fecha(fecha_str):
    if not fecha_str:
        return None
    fecha_str = fecha_str.lower()
    meses = {"enero":1,"febrero":2,"marzo":3,"abril":4,"mayo":5,"junio":6,
             "julio":7,"agosto":8,"septiembre":9,"setiembre":9,"octubre":10,"noviembre":11,"diciembre":12}
    match = re.search(r'(\d{1,2})\s+de\s+([a-z]+)\s+(?:de(?:l)?\s+)?(\d{4})', fecha_str)
    if match:
        try:
            fecha = datetime(int(match.group(3)), meses.get(match.group(2), 0), int(match.group(1)))
            return (datetime.now() - fecha).days
        except:
            return None
    for f in ['%d/%m/%Y','%d-%m-%Y','%Y-%m-%d','%d/%m/%y','%d.%m.%Y']:
        try:
            return (datetime.now() - datetime.strptime(fecha_str, f)).days
        except:
            continue
    return None

# =============================================================================
# CLAUDE – Análisis individual
# =============================================================================

def analizar_con_claude(pdf_bytes):
    client = anthropic.Anthropic(api_key=CLAUDE_API_KEY)
    
    hoy = datetime.now().strftime('%d/%m/%Y')
    anio_actual = datetime.now().year
    mes_actual = datetime.now().strftime('%B')
    anio_pasado = anio_actual - 1
    fecha_hace_30_dias = (datetime.now() - timedelta(days=30)).strftime('%d/%m/%Y')
    fecha_hace_90_dias = (datetime.now() - timedelta(days=90)).strftime('%d/%m/%Y')
    fecha_ejemplo_futura = (datetime.now() + timedelta(days=30)).strftime('%d/%m/%Y')

    prompt = f"""Analizá este documento para apostilla en Cancillería Argentina.

🗓️ CONTEXTO TEMPORAL (actualizado automáticamente):
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
• HOY es: {hoy}
• Año ACTUAL: {anio_actual}
• Mes ACTUAL: {mes_actual}

⚠️ REGLAS SOBRE FECHAS - Lee con atención:

FECHAS VÁLIDAS (NO marcar como problema):
• Cualquier fecha del año {anio_actual} hasta hoy ({hoy})
• Fechas recientes de {anio_pasado} (últimos meses)
• Ejemplo: "{fecha_hace_30_dias}" (hace 30 días) = VÁLIDO ✓
• Ejemplo: "{fecha_hace_90_dias}" (hace 90 días) = VÁLIDO ✓

FECHAS PROBLEMÁTICAS (sí marcar como problema):
• Solo fechas FUTURAS (posteriores a {hoy})
• Ejemplo: "{fecha_ejemplo_futura}" = FUTURO (problemático) ✗

REGLA SIMPLE: 
Si fecha ≤ {hoy} → VÁLIDA, NO marcar problema
Si fecha > {hoy} → FUTURA, marcar problema

NO menciones "{anio_actual}" como algo raro o futuro - ES EL AÑO ACTUAL.

📋 INSTRUCCIONES DE EXTRACCIÓN:

Para calidad_imagen - usá SOLO estas palabras exactas:
• "alta" o "clara" o "nítida" → si se lee bien
• "baja" → si cuesta leer pero se puede
• "borrosa" → si hay desenfoque notable
• "ilegible" → si no se puede leer

Para multiples_firmas:
• Marcá true SOLO si hay firmas de distintas autoridades que generan confusión real sobre cuál es la válida
• Si hay una sola firma clara, marcá false

Para problemas_detectados:
• Listá SOLO problemas DOCUMENTALES que impidan o compliquen la apostilla
• Problemas válidos: firma ausente, imagen ilegible, fecha futura, calidad baja, foto de celular
• NO incluyas la fecha como problema si es de {anio_actual}
• NO incluyas circunstancias del contenido del documento (causa de muerte, tipo de delito, intervención judicial, antecedentes penales registrados, etc.) — eso NO es un problema documental
• Si el documento está formalmente bien, dejá la lista vacía []

Para observacion_redactada:
• Escribí UNA oración clara y profesional que resuma el documento
• Incluí siempre: tipo de documento, titular, fecha, firmante
• Si el documento tiene detalles de contenido interesantes o llamativos, mencionálos naturalmente al final — por ejemplo: causa de muerte inusual, intervención judicial, antecedentes penales registrados, etc.
• Estos detalles son informativos, NO son problemas. El tono debe ser descriptivo y profesional, no alarmista
• Ejemplos de buenas observaciones:
  - "Acta de defunción de José Miguel Sandoval Rojas emitida el 29/01/2026 por Lucrecia Olivieri, con firma digital válida. Fallecimiento por herida de arma de fuego con intervención de UFI 03."
  - "Certificado de antecedentes penales de Juan García emitido el 15/02/{anio_actual}, vigente, con firma digital. Registra antecedentes penales por robo."
  - "Acta de nacimiento de Sofía López emitida el 03/01/{anio_actual} por el Registro Civil, con firma digital de María Rodríguez."

Para titular_documento:
• El nombre completo de la persona a quien pertenece el documento
• Buscá el nombre en TODO el documento, incluso manuscrito o en anotaciones marginales
• En acta de nacimiento: nombre del nacido (ej: "Joel Lautaro Sueldo")
• En antecedente penal: nombre del solicitante
• En título o certificado educativo: nombre del alumno/graduado
• Campo OBLIGATORIO, nunca vacío si el nombre aparece

Para firmantes_visibles:
• Listá TODOS los firmantes que aparecen en CUALQUIER PÁGINA del documento
• Incluí firmantes de sellos ministeriales, no solo de la autoridad educativa
• Ejemplos: ["Prof. Liliana Villena - Rectora", "Maria Alejandra Gutierrez - Legalizaciones GCBA"]
• Si en un sello dice el nombre del firmante, incluilo aunque sea del sello ministerial
• NUNCA dejes esta lista vacía si hay firmas o nombres en el documento

🎓 DOCUMENTOS EDUCATIVOS — Regla de intervención ministerial:
Si el documento es educativo (título, analítico, certificado de materias, diploma, boletín, historia académica, constancia de estudios, etc.):

⚠️ REVISÁ TODAS LAS PÁGINAS DEL PDF ANTES DE RESPONDER — el sello ministerial puede estar en la última página, dorso, o hoja anexa. NO concluyas que no hay sello hasta haber revisado absolutamente todas las páginas.

ORGANISMOS QUE CONSTITUYEN INTERVENCIÓN VÁLIDA:
✅ Ministerio de Educación (nacional o provincial o CABA)
✅ Ministerio de Capital Humano (reemplazó al de Educación a nivel nacional)
✅ Ministerio de Educación e Innovación (CABA)
✅ GOBIERNO DE LA CIUDAD DE BUENOS AIRES — LEGALIZACIONES (sello morado/violeta/púrpura con fecha, firmado por funcionario del Ministerio de Educación GCBA).
   IMPORTANTE: Este sello aparece en la SEGUNDA PÁGINA o hoja separada. Es un rectángulo de tinta morada/violeta que dice "GOBIERNO DE LA CIUDAD DE BUENOS AIRES / LEGALIZACIONES" con una fecha y firma manuscrita. Si ves cualquier rectángulo violeta/morado con texto de GCBA y una firma → es intervención válida. Aunque esté borroso → "intervencion_sello_presente_borroso": true, NO "tiene_intervencion_ministerial": false.
✅ Dirección General de Cultura y Educación (Provincia de Buenos Aires)
✅ DiNIECE, Secretaría de Educación, o cualquier repartición nacional/provincial con competencia en legalización educativa
✅ Ministerio del Interior (SOLO para documentos emitidos ANTES del año 2012 — en esa época era el órgano habilitado para legalizar docs educativos)

❌ ORGANISMOS QUE NO SON INTERVENCIÓN VÁLIDA PARA LEGALIZACIÓN EDUCATIVA:
❌ Ministerio de Relaciones Exteriores y Culto / Cancillería — ese sello es de APOSTILLA, no de legalización educativa. NO cuenta como intervención ministerial educativa.
❌ Consulado General de Chile u otro consulado extranjero — tampoco cuenta.
❌ Cualquier sello de la propia institución educativa (rector, director, secretario, decano).

Para documentos PRE-2012: el ÚNICO organismo válido es el Ministerio del Interior. Si ves sello de Cancillería/RR.EE. en un doc pre-2012 → NO es intervención válida, no marques true.

DOCUMENTOS PRE-2012:
• Si el documento fue emitido ANTES del año 2012, la intervención válida era del Ministerio del Interior.
• NO importa si el documento tiene materias adeudadas, promedios bajos, o cualquier contenido académico — eso NO es un problema documental.
• Si un documento pre-2012 solo tiene firma educativa sin Ministerio del Interior → "tiene_intervencion_ministerial": false
• Si un documento pre-2012 tiene sello del Ministerio del Interior → "tiene_intervencion_ministerial": true
• Si un documento pre-2012 tiene sello del Ministerio del Interior pero está borroso/confuso → null + borroso: true

IMPORTANTE — Múltiples sellos y firmas (documentos complejos):
• Si el documento tiene muchos sellos, firmas superpuestas, o es difícil identificar claramente de qué organismo es cada sello → "multiples_firmas": true, "intervencion_sello_presente_borroso": true
• En esos casos NO asumas que está aprobado — mandá a revisión manual.

IMPORTANTE — Sello borroso o parcialmente legible:
• Si hay un sello que parece ser ministerial pero está borroso, parcialmente legible, poco claro, o con muchos sellos superpuestos → "tiene_intervencion_ministerial": null, "intervencion_sello_presente_borroso": true
• NO marques false si hay un sello visible aunque sea difícil de leer. False es SOLO cuando claramente no hay NINGÚN sello ministerial en ninguna página.

REGLAS:
• Si SOLO tiene firma de la autoridad educativa (rector, decano, director, secretario académico) sin ningún sello ministerial en ninguna página → "tiene_intervencion_ministerial": false
• Si tiene sello/firma ministerial clara en cualquier página → "tiene_intervencion_ministerial": true
• Si el sello está presente pero borroso/parcial/confuso → "tiene_intervencion_ministerial": null + "intervencion_sello_presente_borroso": true
• Si el documento NO es educativo → "tiene_intervencion_ministerial": null

Para es_documento_educativo:
• true si es título, analítico, certificado académico, diploma, boletín de notas, historia académica, certificado de materias, constancia de estudios, etc.
• false para cualquier otro tipo de documento

Para nota_intervencion:
• Si el documento es pre-2012 y necesita Ministerio del Interior, explicalo aquí
• Si el sello está borroso, describí lo que se ve
• En cualquier otro caso, dejá vacío ""

Campos a extraer (JSON válido):
{{
  "tipo_documento": string,
  "titular_documento": string,
  "fecha_emision": string (tal como aparece),
  "anio_documento": number,
  "es_pre_2012": boolean,
  "firmantes_visibles": [strings],
  "cantidad_firmas_visibles": number,
  "multiples_firmas": boolean,
  "sello_ministerio_visible": boolean,
  "sello_claro": boolean,
  "es_documento_educativo": boolean,
  "tiene_intervencion_ministerial": boolean|null,
  "intervencion_sello_presente_borroso": boolean,
  "nota_intervencion": string,
  "calidad_imagen": "alta"|"clara"|"nítida"|"baja"|"borrosa"|"ilegible",
  "es_foto_celular": boolean,
  "problemas_detectados": [strings vacía si todo OK],
  "observacion_redactada": string
}}"""

    message = client.messages.create(
        model="claude-sonnet-4-20250514",
        max_tokens=2000,
        messages=[{"role": "user", "content": [
            {"type": "document", "source": {"type": "base64", "media_type": "application/pdf", "data": pdf_a_base64(pdf_bytes)}},
            {"type": "text", "text": prompt}
        ]}]
    )

    respuesta = message.content[0].text.strip() if message.content else ""
    respuesta = re.sub(r'^```json\n?', '', respuesta)
    respuesta = re.sub(r'\n?```$', '', respuesta)
    # Si Claude envolvió el JSON en texto, extraer solo el objeto
    match_json = re.search(r'\{[\s\S]*\}', respuesta)
    if match_json:
        respuesta = match_json.group(0)
    if not respuesta:
        raise ValueError("Respuesta vacía de Claude. Reintentá o revisá el PDF manualmente.")
    return json.loads(respuesta)

# =============================================================================
# CLAUDE – Análisis de PAR IF + CE (PDF combinado)
# =============================================================================

def analizar_par_if_ce_con_claude(if_bytes, ce_bytes, nombre_if, nombre_ce):
    """
    Combina el IF y el CE en un solo PDF y lo envía a Claude para análisis unificado.
    La firma que importa es la del CE. El IF es el documento original.
    """
    client = anthropic.Anthropic(api_key=CLAUDE_API_KEY)
    
    hoy = datetime.now().strftime('%d/%m/%Y')
    anio_actual = datetime.now().year

    # Combinar ambos PDFs en uno
    pdf_combinado = combinar_pdfs([if_bytes, ce_bytes])

    prompt = f"""Estás analizando un PAR de documentos vinculados para apostilla en Cancillería Argentina.

📂 DOCUMENTO 1 (páginas iniciales): Archivo IF – Es el ACTA o documento original (ej: acta de nacimiento).
📂 DOCUMENTO 2 (páginas siguientes): Archivo CE – Es el CERTIFICADO que avala al IF.

HOY: {hoy} | AÑO ACTUAL: {anio_actual}

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
🔎 TU TAREA PRINCIPAL: Verificar la vinculación IF ↔ CE
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

El CE debe contener en su texto la frase:
"Número/s de documento/s electrónico/s: [número IF]"

El número IF del primer archivo es: {nombre_if}

Verificá si el CE (segundo documento) hace referencia a ese número IF en su texto.

⚠️ IMPORTANTE SOBRE FIRMAS:
• La firma que IMPORTA para apostilla es la del CE (segundo documento), NO la del IF.
• El IF puede tener firma ológrafa (manuscrita) o sellos – eso es NORMAL, no es un problema.
• Evaluá la firma del CE: debe ser digital/electrónica, emitida por GCABA (DGRC).

🖊️ CÓMO EXTRAER EL FIRMANTE DEL CE – Lee con mucha atención:
Los documentos CE de GCABA tienen un bloque de firma que dice:
  "Digitally signed by Comunicaciones Oficiales"
  "Date: YYYY.MM.DD HH:MM:SS"
  
  [Nombre Apellido]         ← ESTE es el firmante_ce que querés
  [Cargo]
  [Organismo]

• "Comunicaciones Oficiales" NO es el firmante. Es el sistema técnico que certifica.
• El firmante real es el NOMBRE HUMANO que aparece DEBAJO del bloque "Digitally signed".
• Ejemplo: si ves "Gonzalo Alvarez / Gerente Operativo / D.G.REG.ESTADO CIVIL..." → firmante_ce = "Gonzalo Alvarez"
• Puede haber dos bloques "Digitally signed" en el CE (uno arriba y uno abajo). En ambos casos el nombre humano aparece debajo. Tomá el primero que encuentres con nombre legible.
• Si no encontrás ningún nombre humano → firmante_ce = "No identificado"

Para calidad_imagen - usá SOLO: "alta", "clara", "nítida", "baja", "borrosa" o "ilegible"

Para titular_documento:
• El nombre completo de la persona del ACTA (IF), ej: "Apolo Luciano Arce Chumbi"
• Buscá en el documento manuscrito o impreso

Para fecha_emision:
• Usá la fecha del CE (no la del IF original), porque el CE es el que tiene vigencia actual

Para observacion_redactada:
• Una sola oración que resuma el par: tipo de acta, titular, si el CE vincula correctamente al IF, y quién firmó el CE.
• Ejemplo: "Acta de nacimiento de Apolo Luciano Arce Chumbi, CE emitido el 20/02/2026, firmado por Gonzalo Alvarez, referencia IF verificada correctamente."

Respondé SOLO JSON válido:
{{
  "tipo_documento": string,
  "titular_documento": string,
  "fecha_emision": string,
  "anio_documento": number,
  "es_pre_2012": boolean,
  "firmantes_visibles": [strings],
  "cantidad_firmas_visibles": number,
  "multiples_firmas": boolean,
  "sello_ministerio_visible": boolean,
  "sello_claro": boolean,
  "calidad_imagen": "alta"|"clara"|"nítida"|"baja"|"borrosa"|"ilegible",
  "es_foto_celular": boolean,
  "ce_referencia_if_correctamente": boolean,
  "numero_if_encontrado_en_ce": string,
  "firmante_ce": string,
  "cargo_firmante_ce": string,
  "problemas_detectados": [],
  "observacion_redactada": string
}}"""

    message = client.messages.create(
        model="claude-sonnet-4-20250514",
        max_tokens=2000,
        messages=[{"role": "user", "content": [
            {"type": "document", "source": {"type": "base64", "media_type": "application/pdf", "data": pdf_a_base64(pdf_combinado)}},
            {"type": "text", "text": prompt}
        ]}]
    )

    respuesta = message.content[0].text.strip()
    respuesta = re.sub(r'^```json\n?', '', respuesta)
    respuesta = re.sub(r'\n?```$', '', respuesta)
    return json.loads(respuesta)

# =============================================================================
# LÓGICA NORMATIVA
# =============================================================================

def evaluar_documento(firma_info, analisis, sello_violeta=False):
    tipo = (analisis.get('tipo_documento') or "").lower()
    estado = "✅ OK"
    accion = "Listo para cargar"
    problemas = []

    calidad = (analisis.get("calidad_imagen") or "").lower()
    problemas_claude = analisis.get("problemas_detectados") or []

    # PENAL: vencimiento 90 días
    if "antecedente" in tipo or "penal" in tipo:
        fecha = analisis.get('fecha_emision')
        if not fecha:
            estado = "⚠️ REVISAR"
            accion = "No se detectó fecha"
            problemas.append("No se pudo leer la fecha de emisión")
        else:
            dias = calcular_dias_desde_fecha(fecha)
            
            if dias is None:
                estado = "⚠️ REVISAR"
                accion = "Fecha no interpretable"
                problemas.append(f"No se pudo interpretar la fecha: {fecha}")
            
            elif dias < 0:
                estado = "⚠️ REVISAR"
                accion = "Fecha posterior a hoy"
                problemas.append(f"Fecha futura detectada: {fecha} (verificar si es error de sistema)")
            
            elif dias > 90:
                estado = "❌ RECHAZAR"
                accion = "Certificado vencido (>90 días)"
                problemas.append(f"Vencido hace {dias} días (máximo: 90)")
            
            else:
                estado = "✅ OK"
                accion = "Certificado vigente"

        if firma_info['tiene_firma'] == False:
            estado = "❌ RECHAZAR"
            accion = "Falta firma digital"
            problemas.append("No se detectó firma digital")

    # Títulos, analíticos y documentos educativos en general
    es_educativo = analisis.get('es_documento_educativo', False)
    tiene_intervencion = analisis.get('tiene_intervencion_ministerial', None)
    sello_borroso = analisis.get('intervencion_sello_presente_borroso', False)
    nota_intervencion = analisis.get('nota_intervencion', '')
    es_pre_2012 = analisis.get('es_pre_2012', False)

    # Detectar por tipo también, como fallback
    palabras_educativas = ["título", "analítico", "certificado académico", "diploma",
                           "historia académica", "boletín", "materias", "carrera",
                           "universidad", "facultad", "escuela", "colegio", "licenciado",
                           "licenciatura", "ingeniero", "ingeniería", "bachiller", "tecnicatura",
                           "constancia de estudios", "certificado de estudios", "estudios parciales"]
    if any(p in tipo for p in palabras_educativas):
        es_educativo = True

    if es_educativo:
        if analisis.get('cantidad_firmas_visibles', 0) == 0:
            if estado == "✅ OK":
                estado = "⚠️ REVISAR"
                accion = "No se detecta firma visible"
            problemas.append("Sin firma visible")

        # Verificar calidad de la detección cuando tiene_intervencion == True
        if tiene_intervencion == True:
            firmantes = [f.lower() for f in (analisis.get('firmantes_visibles') or [])]
            obs = (analisis.get('observacion_redactada') or '').lower()
            nota = (analisis.get('nota_intervencion') or '').lower()
            texto_combinado = ' '.join(firmantes) + ' ' + obs + ' ' + nota

            palabras_ministeriales = [
                'ministerio', 'legalizaciones', 'gcba', 'capital humano',
                'interior', 'dirección general de cultura', 'dgcye', 'diNIECE'
            ]
            palabras_no_validas = [
                'relaciones exteriores', 'cancillería', 'cancilleria',
                'consulado', 'consul adjunto', 'belluomini'
            ]
            # Palabras de encabezado institucional del documento (no sello de legalización)
            solo_encabezado = (
                'ministerio de educaci' in texto_combinado and
                not any(p in texto_combinado for p in ['legalizaciones', 'interior', 'certif', 'sello'])
            )

            tiene_firmante_ministerial = any(p in texto_combinado for p in palabras_ministeriales)
            solo_rree = (
                any(p in texto_combinado for p in palabras_no_validas) and
                'ministerio del interior' not in texto_combinado and
                'legalizaciones' not in texto_combinado
            )

            if solo_rree:
                # RR.EE./Cancillería es apostilla, no legalización educativa
                tiene_intervencion = None
                sello_borroso = True
                problemas.append(
                    "El sello de Cancillería/Ministerio de RR.EE. es de apostilla, "
                    "NO de legalización educativa. "
                    + ("Para documentos pre-2012 se requiere sello del Ministerio del Interior."
                       if es_pre_2012 else
                       "Se requiere sello del Ministerio de Educación o GCBA Legalizaciones.")
                    + " Verificar manualmente."
                )
            elif solo_encabezado and not tiene_firmante_ministerial:
                # Claude leyó el encabezado del documento, no un sello real de legalización
                tiene_intervencion = None
                sello_borroso = True
                problemas.append(
                    "No se identificó firmante de organismo ministerial habilitado. "
                    "El texto 'Ministerio de Educación' detectado parece ser el encabezado "
                    "del certificado, no un sello de legalización. Verificar manualmente."
                )
            elif es_pre_2012 and analisis.get('multiples_firmas'):
                # Pre-2012 con múltiples sellos mezclados → siempre REVISAR
                tiene_intervencion = None
                sello_borroso = True
                problemas.append(
                    "Documento pre-2012 con múltiples sellos y firmas. "
                    "Verificar manualmente que el sello del Ministerio del Interior "
                    "esté presente y sea legible."
                )

        # Sello violeta detectado por análisis de color
        # Se verifica siempre — incluso si Claude dijo true, puede haber confundido
        # el encabezado impreso del documento con un sello real de legalización
        if sello_violeta:
            firmantes = [f.lower() for f in (analisis.get('firmantes_visibles') or [])]
            palabras_edu = ['rector', 'rectora', 'director', 'directora', 'secretario',
                           'secretaria', 'decano', 'decana', 'vicerector', 'vicedirector']
            palabras_min = ['ministerio', 'legalizaciones', 'gcba', 'interior',
                           'capital humano', 'educacion']
            solo_edu = (
                any(p in ' '.join(firmantes) for p in palabras_edu) and
                not any(p in ' '.join(firmantes) for p in palabras_min)
            )
            if tiene_intervencion != True:
                # Claude no confirmó → marcar borroso
                sello_borroso = True
                if "sello violeta" not in " ".join(problemas).lower():
                    problemas.append(
                        "Sello violeta/morado detectado en el PDF (posible sello ministerial GCBA). "
                        "No fue confirmado automáticamente — verificar manualmente."
                    )
            elif solo_edu:
                # Claude dijo true pero solo hay firmantes educativos → encabezado confundido con sello
                tiene_intervencion = None
                sello_borroso = True
                if "sello violeta" not in " ".join(problemas).lower():
                    problemas.append(
                        "Sello violeta/morado detectado pero solo se identificaron autoridades educativas "
                        "como firmantes. Posible confusión entre encabezado del documento y sello ministerial "
                        "real — verificar manualmente que exista intervención de GCBA Legalizaciones."
                    )

        # Sello presente pero borroso → REVISAR (no rechazar)
        if sello_borroso or tiene_intervencion is None:
            if estado == "✅ OK":
                estado = "⚠️ REVISAR"
                accion = "Verificar sello ministerial manualmente"
            if sello_borroso:
                msg = "Sello ministerial presente pero borroso o parcialmente legible — verificar manualmente."
                if nota_intervencion:
                    msg += f" {nota_intervencion}"
                problemas.append(msg)
            elif es_pre_2012 and nota_intervencion:
                estado = "⚠️ REVISAR"
                accion = "Requiere verificación de intervención ministerial"
                problemas.append(nota_intervencion)

        # Sin intervención detectada:
        # Si calidad buena (Claude vio bien y no hay sello) → RECHAZAR
        # Si calidad mala/borrosa (imagen comprimida, sello puede ser invisible) → REVISAR
        elif tiene_intervencion == False:
            calidad_ok = calidad in ["alta", "clara", "nítida"]

            if analisis.get('sello_ministerio_visible'):
                # Claude vio algo pero no lo confirmó — siempre REVISAR
                estado = "⚠️ REVISAR"
                accion = "Sello ministerial detectado — confirmar manualmente"
                problemas.append(
                    "Se detectó un sello que podría ser ministerial pero no fue confirmado. "
                    "Verificar manualmente si corresponde al Ministerio de Educación o habilitado."
                )
            elif calidad_ok:
                # Imagen nítida y Claude no vio nada → RECHAZAR con confianza
                estado = "❌ RECHAZAR"
                if es_pre_2012:
                    accion = "Falta intervención del Ministerio del Interior"
                    problemas.append(
                        "Documento educativo anterior a 2012. "
                        "Requiere intervención del Ministerio del Interior. "
                        "La firma de la autoridad educativa no es suficiente para apostillar."
                    )
                else:
                    accion = "Falta intervención del Ministerio de Educación"
                    problemas.append(
                        "El documento solo tiene firma de la autoridad educativa. "
                        "Para apostillar es obligatoria la intervención del Ministerio de "
                        "Educación, Capital Humano, o GCBA Legalizaciones."
                    )
            else:
                # Imagen baja/borrosa — no confiar en la no-detección → REVISAR
                estado = "⚠️ REVISAR"
                if es_pre_2012:
                    accion = "Verificar sello del Ministerio del Interior"
                    problemas.append(
                        "Documento educativo anterior a 2012. Requiere sello del Ministerio del Interior. "
                        "No se detectó automáticamente — verificar manualmente (imagen de baja calidad, "
                        "el sello puede estar borroso o invisible por compresión)."
                    )
                else:
                    accion = "Verificar sello del Ministerio de Educación"
                    problemas.append(
                        "No se detectó sello ministerial — imagen de baja calidad o comprimida. "
                        "Verificar manualmente que tenga intervención del Ministerio de Educación, "
                        "Capital Humano, o GCBA Legalizaciones (sellos violetas pueden no detectarse)."
                    )
    elif "título" in tipo or "analítico" in tipo:
        if analisis.get('cantidad_firmas_visibles', 0) == 0:
            if estado == "✅ OK":
                estado = "⚠️ REVISAR"
                accion = "No se detecta firma visible"
            problemas.append("Sin firma visible")

    # Múltiples firmas — para docs físicos se basa en lo que Claude detectó
    if analisis.get('multiples_firmas'):
        if estado == "✅ OK":
            estado = "⚠️ REVISAR"
            accion = "Múltiples sellos/firmas — verificar manualmente"
        problemas.append("Múltiples firmas o sellos detectados — verificar cuál corresponde")

    # Calidad
    if calidad == "ilegible":
        estado = "❌ RECHAZAR"
        accion = "Imagen ilegible"
        problemas.append("Imagen ilegible")
    elif calidad in ["baja", "borrosa"]:
        if estado == "✅ OK":
            estado = "⚠️ REVISAR"
            accion = "Calidad de imagen insuficiente"
        problemas.append(f"Calidad de imagen: {calidad}")

    # Problemas detectados por Claude (filtrar falsos positivos de fecha)
    problemas_filtrados = []
    for p in problemas_claude:
        p_lower = p.lower()
        anio_actual = str(datetime.now().year)
        if anio_actual not in p_lower and "fecha futura" not in p_lower and "fecha posterior" not in p_lower:
            problemas_filtrados.append(p)
    
    if problemas_filtrados:
        if estado == "✅ OK":
            estado = "⚠️ REVISAR"
            accion = "Revisar problemas detectados"
        for p in problemas_filtrados:
            problemas.append(p)

    # Foto de celular
    if analisis.get("es_foto_celular"):
        if estado == "✅ OK":
            estado = "⚠️ REVISAR"
            accion = "Documento fotografiado con celular"
        problemas.append("Documento fotografiado con celular")

    return estado, accion, problemas

def evaluar_par_if_ce(firma_info_ce, analisis_par):
    """
    Evalúa el resultado del análisis de un par IF+CE.
    La firma que importa es la del CE.
    """
    estado = "✅ OK"
    accion = "Par IF+CE válido – Listo para cargar"
    problemas = []

    # Verificación principal: ¿el CE hace referencia al IF?
    if not analisis_par.get("ce_referencia_if_correctamente"):
        estado = "❌ RECHAZAR"
        accion = "El CE no referencia al IF correspondiente"
        problemas.append("El CE no contiene el número IF correcto en su texto")

    # Firma del CE
    if firma_info_ce['tiene_firma'] == False:
        estado = "❌ RECHAZAR"
        accion = "CE sin firma digital"
        problemas.append("El CE no tiene firma digital válida")
    elif firma_info_ce['tiene_firma'] is None:
        if estado == "✅ OK":
            estado = "⚠️ REVISAR"
            accion = "Firma del CE no detectada automáticamente"
        problemas.append("No se pudo verificar firma digital del CE automáticamente")

    # Calidad
    calidad = (analisis_par.get("calidad_imagen") or "").lower()
    if calidad == "ilegible":
        estado = "❌ RECHAZAR"
        accion = "Imagen ilegible"
        problemas.append("Imagen ilegible")
    elif calidad in ["baja", "borrosa"]:
        if estado == "✅ OK":
            estado = "⚠️ REVISAR"
            accion = "Calidad de imagen insuficiente"
        problemas.append(f"Calidad de imagen: {calidad}")

    # Problemas detectados por Claude (filtrar falsos positivos)
    problemas_claude = analisis_par.get("problemas_detectados") or []
    anio_actual = str(datetime.now().year)
    for p in problemas_claude:
        p_lower = p.lower()
        if anio_actual not in p_lower and "fecha futura" not in p_lower and "fecha posterior" not in p_lower:
            if estado == "✅ OK":
                estado = "⚠️ REVISAR"
                accion = "Revisar problemas detectados"
            problemas.append(p)

    return estado, accion, problemas

def generar_observacion(analisis, problemas):
    obs_base = analisis.get("observacion_redactada") or analisis.get("observaciones") or ""
    if problemas:
        extra = "; ".join(problemas)
        if obs_base:
            return f"{obs_base.strip()} — {extra}"
        return extra
    return obs_base.strip()

# =============================================================================
# EXCEL EN MEMORIA
# =============================================================================

def generar_excel(df):
    output = io.BytesIO()
    df.to_excel(output, index=False, engine="openpyxl")
    output.seek(0)
    wb = load_workbook(output)
    ws = wb.active

    verde = PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid")
    amarillo = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")
    rojo = PatternFill(start_color="FDECEA", end_color="FDECEA", fill_type="solid")
    gris = PatternFill(start_color="ECECEC", end_color="ECECEC", fill_type="solid")

    for cell in ws[1]:
        cell.fill = gris
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    col_estado = next((i for i, c in enumerate(ws[1], 1) if c.value == "Estado"), None)

    for row in ws.iter_rows(min_row=2):
        estado = row[col_estado - 1].value if col_estado else ""
        fill = verde if "OK" in str(estado) else amarillo if "REVISAR" in str(estado) else rojo if "RECHAZAR" in str(estado) else None
        if fill:
            for cell in row:
                cell.fill = fill

    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 60)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    final = io.BytesIO()
    wb.save(final)
    final.seek(0)
    return final

# =============================================================================
# INTERFAZ
# =============================================================================

with tab_revision:
    # Texto encima del uploader
    st.markdown("""
<div style="text-align: center; margin-bottom: 0.75rem; padding-top: 0.5rem;">
    <p style="font-size: 0.92rem; font-weight: 500; color: #F0F0F2; margin: 0 0 0.2rem 0; letter-spacing: -0.01em;">
        Arrastrá los PDFs acá
    </p>
    <p style="font-size: 0.78rem; color: #7A7A85; margin: 0;">
        Podés subir varios archivos a la vez — IF y CE juntos
    </p>
</div>
""", unsafe_allow_html=True)

    # Ícono encima del uploader — elemento independiente, no ligado al padding del dropzone
    st.markdown("""
<div style="
    display: flex; flex-direction: column; align-items: center;
    margin-bottom: -5rem; position: relative; z-index: 1; pointer-events: none;
    padding-top: 1.6rem;
">
    <div style="
        width: 52px; height: 52px; border-radius: 13px;
        background: rgba(79,142,247,0.1); border: 1px solid rgba(79,142,247,0.2);
        display: flex; align-items: center; justify-content: center;
    ">
        <svg width="22" height="22" viewBox="0 0 24 24" fill="none" xmlns="http://www.w3.org/2000/svg">
            <path d="M12 15V3M12 3L8 7M12 3L16 7" stroke="#4F8EF7" stroke-width="1.75" stroke-linecap="round" stroke-linejoin="round"/>
            <path d="M3 15V18C3 19.1046 3.89543 20 5 20H19C20.1046 20 21 19.1046 21 18V15" stroke="#4F8EF7" stroke-width="1.75" stroke-linecap="round"/>
        </svg>
    </div>
</div>
""", unsafe_allow_html=True)

    archivos = st.file_uploader("PDFs", type=["pdf"], accept_multiple_files=True, label_visibility="collapsed")

    # Renombrar el botón nativo "Browse files" → "Seleccionar archivos"
    st.markdown("""
<script>
(function renameBtn() {
    const btn = document.querySelector('[data-testid="stFileUploaderDropzone"] button');
    if (btn && btn.innerText.trim() === "Browse files") {
        btn.innerText = "Seleccionar archivos";
    } else {
        setTimeout(renameBtn, 100);
    }
})();
</script>
""", unsafe_allow_html=True)

    if not archivos:
        st.markdown("""
<div style="
    margin-top: 2.5rem;
    display: flex;
    flex-direction: column;
    align-items: center;
    text-align: center;
    padding: 2.5rem 1rem;
    opacity: 0.85;
">
    <div style="
        width: 52px; height: 52px;
        border: 1px solid rgba(255,255,255,0.07);
        border-radius: 14px;
        display: flex; align-items: center; justify-content: center;
        margin-bottom: 1.1rem;
        background: #18181B;
    ">
        <svg width="22" height="22" viewBox="0 0 24 24" fill="none" xmlns="http://www.w3.org/2000/svg">
            <path d="M3 7C3 5.89543 3.89543 5 5 5H9.58579C9.851 5 10.1054 5.10536 10.2929 5.29289L11.7071 6.70711C11.8946 6.89464 12.149 7 12.4142 7H19C20.1046 7 21 7.89543 21 9V17C21 18.1046 20.1046 19 19 19H5C3.89543 19 3 18.1046 3 17V7Z"
              stroke="#7A7A85" stroke-width="1.6" stroke-linecap="round" stroke-linejoin="round"/>
        </svg>
    </div>
    <p style="font-size: 0.95rem; font-weight: 500; color: #F0F0F2; margin: 0 0 0.35rem 0; letter-spacing: -0.01em;">
        Ningún documento cargado
    </p>
    <p style="font-size: 0.82rem; color: #7A7A85; margin: 0 0 1.8rem 0; max-width: 320px; line-height: 1.6;">
        Subí los PDFs arriba para comenzar. El sistema detecta automáticamente si son IF, CE o documentos individuales.
    </p>
    <div style="display: flex; gap: 1.5rem; justify-content: center; flex-wrap: wrap;">
        <div style="
            background: #18181B; border: 1px solid rgba(255,255,255,0.07);
            border-radius: 10px; padding: 0.75rem 1.1rem; min-width: 130px;
        ">
            <p style="font-size: 9px; font-weight: 600; letter-spacing: 0.1em; text-transform: uppercase; color: #7A7A85; margin: 0 0 0.3rem 0;">Acepta</p>
            <p style="font-size: 0.82rem; color: #B0B0BA; margin: 0; font-weight: 500;">Solo archivos PDF</p>
        </div>
        <div style="
            background: #18181B; border: 1px solid rgba(255,255,255,0.07);
            border-radius: 10px; padding: 0.75rem 1.1rem; min-width: 130px;
        ">
            <p style="font-size: 9px; font-weight: 600; letter-spacing: 0.1em; text-transform: uppercase; color: #7A7A85; margin: 0 0 0.3rem 0;">Cantidad</p>
            <p style="font-size: 0.82rem; color: #B0B0BA; margin: 0; font-weight: 500;">Múltiples a la vez</p>
        </div>
        <div style="
            background: #18181B; border: 1px solid rgba(255,255,255,0.07);
            border-radius: 10px; padding: 0.75rem 1.1rem; min-width: 130px;
        ">
            <p style="font-size: 9px; font-weight: 600; letter-spacing: 0.1em; text-transform: uppercase; color: #7A7A85; margin: 0 0 0.3rem 0;">Clasifica</p>
            <p style="font-size: 0.82rem; color: #B0B0BA; margin: 0; font-weight: 500;">IF · CE · Individual</p>
        </div>
    </div>
</div>
""", unsafe_allow_html=True)

    if archivos:
        # ── Pre-clasificación para el contador ───────────────────────────────
        _preview_if = {}
        _preview_ce = {}
        _preview_otros = 0

        for _a in archivos:
            _bytes = _a.read()
            _a.seek(0)  # reset para que el análisis pueda releerlo después
            _tipo, _clave, _ = detectar_tipo_por_contenido(_bytes, _a.name)
            if _tipo == "IF":
                _preview_if[_clave] = True
            elif _tipo == "CE":
                _preview_ce[_a.name] = _clave
            else:
                _preview_otros += 1

        # Calcular pares y huérfanos
        _pares_count = sum(1 for clave in _preview_ce.values() if clave in _preview_if)
        _huerfanos_ce = sum(1 for clave in _preview_ce.values() if clave not in _preview_if)
        _huerfanos_if = sum(1 for clave in _preview_if if clave not in _preview_ce.values())
        _individuales = _preview_otros + _huerfanos_ce + _huerfanos_if

        # ── Tarjeta de resumen pre-análisis ──────────────────────────────────
        partes_contador = []
        if _pares_count:
            partes_contador.append(f'<span style="color:#4F8EF7; font-weight:600;">{_pares_count} par{"es" if _pares_count != 1 else ""} IF+CE</span>')
        if _individuales:
            partes_contador.append(f'<span style="color:#B0B0BA; font-weight:500;">{_individuales} individual{"es" if _individuales != 1 else ""}</span>')

        separador = '<span style="color:#3A3A3C; margin: 0 0.4rem;">·</span>'
        resumen_html = separador.join(partes_contador) if partes_contador else ""

        st.markdown(f"""
<div style="
    display: flex; align-items: center; justify-content: space-between;
    background: #18181B; border: 1px solid rgba(255,255,255,0.07);
    border-radius: 10px; padding: 0.75rem 1.1rem; margin: 0.8rem 0 0.6rem 0;
">
    <div style="display:flex; align-items:center; gap: 0.6rem;">
        <svg width="15" height="15" viewBox="0 0 24 24" fill="none" xmlns="http://www.w3.org/2000/svg">
            <path d="M9 12h6M9 16h6M7 4H5a2 2 0 00-2 2v14a2 2 0 002 2h14a2 2 0 002-2V6a2 2 0 00-2-2h-2M9 4a2 2 0 002 2h2a2 2 0 002-2M9 4a2 2 0 012-2h2a2 2 0 012 2"
              stroke="#7A7A85" stroke-width="1.75" stroke-linecap="round" stroke-linejoin="round"/>
        </svg>
        <span style="font-size:0.82rem; color:#7A7A85;">{len(archivos)} archivo{"s" if len(archivos) != 1 else ""} detectado{"s" if len(archivos) != 1 else ""}</span>
    </div>
    <div style="font-size:0.82rem;">{resumen_html}</div>
</div>
""", unsafe_allow_html=True)

        if st.button("Analizar documentos", type="primary"):
            resultados = []
            barra = st.progress(0)
            estado_texto = st.empty()

            # ─────────────────────────────────────────────────────────────────
            # PASO 1: Clasificar archivos en IF, CE y OTROS leyendo el CONTENIDO
            # (independiente del nombre del archivo)
            # ─────────────────────────────────────────────────────────────────
            archivos_if = {}   # clave_if (año, numero) → {"archivo": ..., "bytes": ..., "nombre": ...}
            archivos_ce = {}   # nombre → {"archivo": ..., "bytes": ..., "clave_if_ref": ...}
            archivos_otros = []

            # Panel de debug expandible
            with st.expander("Clasificación de archivos — expandí si hay problemas de emparejamiento", expanded=False):
                debug_placeholder = st.empty()
                debug_rows = []

            for archivo in archivos:
                pdf_bytes = archivo.read()
                tipo, clave, texto_extraido = detectar_tipo_por_contenido(pdf_bytes, archivo.name)

                # Info de debug
                preview = texto_extraido[:300].replace("\n", " ") if texto_extraido else "(sin texto extraíble)"
                clave_str = f"IF-{clave[0]}-{clave[1]}" if clave else "—"
                debug_rows.append({
                    "Archivo": archivo.name,
                    "Clasificado como": tipo,
                    "Clave IF detectada": clave_str,
                    "Texto extraído (primeros 300 chars)": preview
                })
                debug_placeholder.dataframe(pd.DataFrame(debug_rows), use_container_width=True)

                if tipo == "IF":
                    archivos_if[clave] = {"archivo": archivo, "bytes": pdf_bytes, "nombre": archivo.name}

                elif tipo == "CE":
                    archivos_ce[archivo.name] = {
                        "archivo": archivo,
                        "bytes": pdf_bytes,
                        "clave_if_ref": clave,
                        "nombre": archivo.name
                    }
                else:
                    archivos_otros.append({"archivo": archivo, "bytes": pdf_bytes})

            # ─────────────────────────────────────────────────────────────────
            # PASO 2: Emparejar CE ↔ IF — solo por coincidencia EXACTA de número
            # No hay fallbacks ciegos: si los números no coinciden, no se emparejan
            # ─────────────────────────────────────────────────────────────────
            pares = []
            if_usados = set()
            ce_usados = set()

            for ce_nombre, ce_data in archivos_ce.items():
                clave_ref = ce_data["clave_if_ref"]

                if clave_ref and clave_ref in archivos_if:
                    # Coincidencia exacta: el CE referencia este IF por número
                    pares.append({
                        "if": archivos_if[clave_ref],
                        "ce": ce_data
                    })
                    if_usados.add(clave_ref)
                    ce_usados.add(ce_nombre)
                else:
                    # Sin coincidencia: CE huérfano (falta el IF o los números no coinciden)
                    ce_usados.add(ce_nombre)
                    archivos_otros.append({
                        "archivo": ce_data["archivo"],
                        "bytes": ce_data["bytes"],
                        "advertencia_ce_sin_if": True,
                        "clave_if_ref": clave_ref
                    })

            # IFs que no fueron referenciados por ningún CE
            for clave_if, if_data in archivos_if.items():
                if clave_if not in if_usados:
                    archivos_otros.append({
                        "archivo": if_data["archivo"],
                        "bytes": if_data["bytes"],
                        "advertencia_if_sin_ce": True
                    })

            total_tareas = len(pares) + len(archivos_otros)
            tarea_actual = 0

            # ─────────────────────────────────────────────────────────────────
            # PASO 3: Procesar PARES IF+CE
            # ─────────────────────────────────────────────────────────────────
            for par in pares:
                if_data = par["if"]
                ce_data = par["ce"]
                nombre_display = f"{if_data['nombre']} + {ce_data['nombre']}"
                estado_texto.text(f"Analizando par: {nombre_display}...")
                barra.progress(tarea_actual / total_tareas if total_tareas else 1)

                try:
                    # Firma: solo la del CE importa
                    firma_info_ce = verificar_firma_digital(ce_data["bytes"])

                    # Análisis conjunto con Claude (PDF combinado)
                    analisis_par = analizar_par_if_ce_con_claude(
                        if_data["bytes"],
                        ce_data["bytes"],
                        if_data["nombre"],
                        ce_data["nombre"]
                    )

                    estado, accion, problemas = evaluar_par_if_ce(firma_info_ce, analisis_par)
                    observacion = generar_observacion(analisis_par, problemas)

                    tiene_firma = firma_info_ce["tiene_firma"]
                    firma_texto = "SÍ" if tiene_firma else ("NO" if tiene_firma == False else "NO DETECTADA")

                    firmante_ce = analisis_par.get("firmante_ce", "") or "No identificado"
                    cargo_ce = analisis_par.get("cargo_firmante_ce", "") or ""
                    firmante_display = f"{firmante_ce} ({cargo_ce})" if cargo_ce else firmante_ce

                    resultados.append({
                        "Archivo": nombre_display,
                        "Tipo trámite": "📎 Par IF+CE",
                        "Titular": analisis_par.get("titular_documento"),
                        "Tipo": analisis_par.get("tipo_documento"),
                        "Fecha CE": analisis_par.get("fecha_emision"),
                        "CE referencia IF": "✅ SÍ" if analisis_par.get("ce_referencia_if_correctamente") else "❌ NO",
                        "IF encontrado en CE": analisis_par.get("numero_if_encontrado_en_ce", ""),
                        "Firmante CE": firmante_display,
                        "Firma Digital CE": firma_texto,
                        "Firmantes Certificado": ", ".join(firma_info_ce["firmantes"]),
                        "Estado": estado,
                        "Acción": accion,
                        "Observaciones": observacion
                    })

                except Exception as e:
                    resultados.append({
                        "Archivo": nombre_display,
                        "Tipo trámite": "📎 Par IF+CE",
                        "Titular": "",
                        "Tipo": "",
                        "Fecha CE": "",
                        "CE referencia IF": "",
                        "IF encontrado en CE": "",
                        "Firmante CE": "",
                        "Firma Digital CE": "",
                        "Firmantes Certificado": "",
                        "Estado": "⚠️ REVISAR",
                        "Acción": "Error de análisis",
                        "Observaciones": f"Error: {str(e)}"
                    })

                tarea_actual += 1

            # ─────────────────────────────────────────────────────────────────
            # PASO 4: Procesar archivos individuales (OTROS, IF sin CE, CE sin IF)
            # ─────────────────────────────────────────────────────────────────
            for item in archivos_otros:
                archivo = item["archivo"]
                pdf_bytes = item["bytes"]
                estado_texto.text(f"Analizando {archivo.name}...")
                barra.progress(tarea_actual / total_tareas if total_tareas else 1)

                advertencia_extra = ""
                if item.get("advertencia_if_sin_ce"):
                    advertencia_extra = "⚠️ IF sin CE correspondiente cargado"
                elif item.get("advertencia_ce_sin_if"):
                    clave_ref = item.get("clave_if_ref")
                    ref_str = f"IF-{clave_ref[0]}-{clave_ref[1]}" if clave_ref else "desconocido"
                    advertencia_extra = f"⚠️ CE sin IF correspondiente (busca: {ref_str})"

                try:
                    firma_info = verificar_firma_digital(pdf_bytes)
                    analisis = analizar_con_claude(pdf_bytes)
                    sello_violeta = detectar_sello_violeta(pdf_bytes)
                    estado, accion, problemas = evaluar_documento(firma_info, analisis, sello_violeta)

                    if advertencia_extra:
                        problemas.append(advertencia_extra)
                        if estado == "✅ OK":
                            estado = "⚠️ REVISAR"
                            accion = advertencia_extra

                    observacion = generar_observacion(analisis, problemas)

                    tiene_firma = firma_info["tiene_firma"]
                    firma_texto = "SÍ" if tiene_firma else ("NO" if tiene_firma == False else "NO DETECTADA")

                    resultados.append({
                        "Archivo": archivo.name,
                        "Tipo trámite": "📄 Individual",
                        "Titular": analisis.get("titular_documento"),
                        "Tipo": analisis.get("tipo_documento"),
                        "Fecha CE": analisis.get("fecha_emision"),
                        "CE referencia IF": "—",
                        "IF encontrado en CE": "—",
                        "Firmante CE": "—",
                        "Firma Digital CE": firma_texto,
                        "Firmantes Certificado": ", ".join(
                            # Para docs educativos: priorizar firmante ministerial educativo
                            # Excluir Cancillería/RR.EE. que es apostilla, no legalización
                            (lambda fv, fi: (
                                [f for f in fv if any(p in f.lower() for p in
                                    ["ministerio de educacion","ministerio educacion","capital humano",
                                     "legalizaciones","gcba","ministerio del interior","interior y trans"])
                                and not any(x in f.lower() for x in
                                    ["relaciones exteriores","cancillería","cancilleria","consulado"])]
                                or [f for f in fv if any(p in f.lower() for p in
                                    ["ministerio","legalizaciones","gcba","interior"])
                                and not any(x in f.lower() for x in
                                    ["relaciones exteriores","cancillería","cancilleria"])]
                                or fv or fi or []
                            ))(
                                analisis.get("firmantes_visibles") or [],
                                firma_info["firmantes"] or []
                            )
                        ),
                        "Estado": estado,
                        "Acción": accion,
                        "Observaciones": observacion
                    })

                except Exception as e:
                    resultados.append({
                        "Archivo": archivo.name,
                        "Tipo trámite": "📄 Individual",
                        "Titular": "",
                        "Tipo": "",
                        "Fecha CE": "",
                        "CE referencia IF": "—",
                        "IF encontrado en CE": "—",
                        "Firmante CE": "—",
                        "Firma Digital CE": "",
                        "Firmantes Certificado": "",
                        "Estado": "⚠️ REVISAR",
                        "Acción": "Error de análisis",
                        "Observaciones": f"Error: {str(e)}"
                    })

                tarea_actual += 1

            barra.progress(1.0)
            estado_texto.text("✓ Análisis completado.")

            # Resumen de pares detectados
            if pares:
                st.success(f"{len(pares)} par(es) IF+CE vinculados correctamente.")

            df = pd.DataFrame(resultados)

            # ── MÉTRICAS ─────────────────────────────────────────────────────
            st.markdown('<div class="section-divider"></div>', unsafe_allow_html=True)
            total = len(df)
            ok = len(df[df["Estado"].str.contains("OK", na=False)])
            revisar = len(df[df["Estado"].str.contains("REVISAR", na=False)])
            rechazar = len(df[df["Estado"].str.contains("RECHAZAR", na=False)])

            col1, col2, col3, col4 = st.columns(4)
            col1.metric("Total analizados", total)
            col2.metric("Aprobados", ok)
            col3.metric("A revisar", revisar)
            col4.metric("Rechazados", rechazar)
            st.markdown('<div class="section-divider"></div>', unsafe_allow_html=True)

            # ── TABLA ─────────────────────────────────────────────────────────
            st.markdown('<p class="results-label">Resultados detallados</p>', unsafe_allow_html=True)

            def colorear_filas(row):
                estado = str(row.get("Estado", ""))
                if "OK" in estado:
                    color = "background-color: rgba(52,199,89,0.08); color: #E0E0E0;"
                elif "REVISAR" in estado:
                    color = "background-color: rgba(255,159,10,0.08); color: #E0E0E0;"
                elif "RECHAZAR" in estado:
                    color = "background-color: rgba(255,69,58,0.08); color: #E0E0E0;"
                else:
                    color = "color: #E0E0E0;"
                return [color] * len(row)

            def colorear_celda_estado(val):
                val = str(val)
                if "OK" in val:
                    return "color: #34C759; font-weight: 600;"
                elif "REVISAR" in val:
                    return "color: #FF9F0A; font-weight: 600;"
                elif "RECHAZAR" in val:
                    return "color: #FF453A; font-weight: 600;"
                return ""

            df_styled = (
                df.style
                .apply(colorear_filas, axis=1)
                .map(colorear_celda_estado, subset=["Estado"])
            )

            st.dataframe(df_styled, use_container_width=True)

            st.markdown('<div style="margin-top:1rem;"></div>', unsafe_allow_html=True)
            st.download_button(
                label="Descargar Excel",
                data=generar_excel(df),
                file_name="revision_apostillas.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )


st.markdown('<div class="footer">Revisor de Apostillas · Automatización documental con IA · Leandro Spinelli · 2026 · v4.0</div>', unsafe_allow_html=True)